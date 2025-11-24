"""
Serviço para geração de relatórios PDF e Excel
"""
from io import BytesIO
from datetime import datetime
from typing import List, Optional, Dict
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.demanda import Demanda
from app.models.cliente import Cliente
from app.models.secretaria import Secretaria
from app.models.tipo_demanda import TipoDemanda
from app.models.prioridade import Prioridade
from app.models.user import User


class RelatorioService:
    """
    Gerador de relatórios PDF e Excel
    """
    
    def gerar_pdf(
        self,
        demandas: List[Demanda],
        filtros: Dict[str, Optional[str]],
        usuario: User
    ) -> BytesIO:
        """
        Gerar relatório PDF
        
        Args:
            demandas: Lista de demandas
            filtros: Filtros aplicados (para mostrar no relatório)
            usuario: Usuário que gerou o relatório
        
        Returns:
            BytesIO com conteúdo do PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#3B82F6'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Título
        title = Paragraph("Relatório de Demandas", title_style)
        elements.append(title)
        elements.append(Spacer(1, 10))
        
        # Data de geração
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        date_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        elements.append(Paragraph(date_text, date_style))
        elements.append(Paragraph(f"Gerado por: {usuario.nome_completo or usuario.username}", date_style))
        elements.append(Spacer(1, 20))
        
        # Resumo
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("<b>Resumo</b>", summary_style))
        
        # Estatísticas rápidas
        total = len(demandas)
        abertas = sum(1 for d in demandas if d.status.value == 'aberta')
        em_andamento = sum(1 for d in demandas if d.status.value == 'em_andamento')
        concluidas = sum(1 for d in demandas if d.status.value == 'concluida')
        canceladas = sum(1 for d in demandas if d.status.value == 'cancelada')
        
        summary_data = [
            ['Total de Demandas', str(total)],
            ['Abertas', str(abertas)],
            ['Em Andamento', str(em_andamento)],
            ['Concluídas', str(concluidas)],
            ['Canceladas', str(canceladas)]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Filtros aplicados
        filtros_aplicados = {k: v for k, v in filtros.items() if v is not None}
        if filtros_aplicados:
            elements.append(Paragraph("<b>Filtros Aplicados</b>", summary_style))
            filtros_text = []
            for key, value in filtros_aplicados.items():
                filtros_text.append(f"• {key.replace('_', ' ').title()}: {value}")
            elements.append(Paragraph("<br/>".join(filtros_text), styles['Normal']))
            elements.append(Spacer(1, 20))
        
        # Tabela de demandas
        if demandas:
            elements.append(Paragraph("<b>Lista de Demandas</b>", summary_style))
            elements.append(Spacer(1, 10))
            
            # Cabeçalho da tabela
            headers = ['ID', 'Nome', 'Tipo', 'Prioridade', 'Status', 'Prazo', 'Cliente']
            data = [headers]
            
            # Dados das demandas
            for demanda in demandas:
                # Obter relacionamentos
                tipo_nome = demanda.tipo_demanda.nome if demanda.tipo_demanda else 'N/A'
                prioridade_nome = demanda.prioridade.nome if demanda.prioridade else 'N/A'
                status_nome = demanda.status.value.replace('_', ' ').title()
                prazo = demanda.prazo_final.strftime('%d/%m/%Y') if demanda.prazo_final else 'N/A'
                
                # Obter cliente via secretaria
                cliente_nome = 'N/A'
                if demanda.secretaria and demanda.secretaria.cliente:
                    cliente_nome = demanda.secretaria.cliente.nome
                
                # Truncar nome se muito longo
                nome = demanda.nome[:40] + '...' if len(demanda.nome) > 40 else demanda.nome
                
                data.append([
                    demanda.id[:8] + '...',  # Primeiros 8 caracteres do UUID
                    nome,
                    tipo_nome,
                    prioridade_nome,
                    status_nome,
                    prazo,
                    cliente_nome[:30] if len(cliente_nome) > 30 else cliente_nome
                ])
            
            # Criar tabela
            table = Table(data, colWidths=[0.8*inch, 2.2*inch, 1*inch, 0.9*inch, 1*inch, 0.9*inch, 1.2*inch])
            table.setStyle(TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                
                # Dados
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID centralizado
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
        else:
            elements.append(Paragraph("<i>Nenhuma demanda encontrada com os filtros aplicados.</i>", styles['Normal']))
        
        # Rodapé
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"DeBrief Sistema - Página 1", footer_style))
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def gerar_excel(
        self,
        demandas: List[Demanda],
        filtros: Dict[str, Optional[str]],
        usuario: User
    ) -> BytesIO:
        """
        Gerar relatório Excel
        
        Args:
            demandas: Lista de demandas
            filtros: Filtros aplicados
            usuario: Usuário que gerou o relatório
        
        Returns:
            BytesIO com conteúdo do Excel
        """
        buffer = BytesIO()
        wb = Workbook()
        
        # ========== ABA: RESUMO ==========
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        # Título
        ws_resumo['A1'] = 'Relatório de Demandas'
        ws_resumo['A1'].font = Font(size=16, bold=True, color="3B82F6")
        ws_resumo.merge_cells('A1:B1')
        
        # Data de geração
        ws_resumo['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        ws_resumo['A2'].font = Font(size=10, italic=True)
        ws_resumo['A3'] = f"Gerado por: {usuario.nome_completo or usuario.username}"
        ws_resumo['A3'].font = Font(size=10, italic=True)
        
        # Estatísticas
        ws_resumo['A5'] = 'Total de Demandas:'
        ws_resumo['B5'] = len(demandas)
        ws_resumo['B5'].font = Font(bold=True)
        
        abertas = sum(1 for d in demandas if d.status.value == 'aberta')
        em_andamento = sum(1 for d in demandas if d.status.value == 'em_andamento')
        concluidas = sum(1 for d in demandas if d.status.value == 'concluida')
        canceladas = sum(1 for d in demandas if d.status.value == 'cancelada')
        
        ws_resumo['A6'] = 'Abertas:'
        ws_resumo['B6'] = abertas
        ws_resumo['A7'] = 'Em Andamento:'
        ws_resumo['B7'] = em_andamento
        ws_resumo['A8'] = 'Concluídas:'
        ws_resumo['B8'] = concluidas
        ws_resumo['A9'] = 'Canceladas:'
        ws_resumo['B9'] = canceladas
        
        # Filtros aplicados
        filtros_aplicados = {k: v for k, v in filtros.items() if v is not None}
        if filtros_aplicados:
            ws_resumo['A11'] = 'Filtros Aplicados:'
            ws_resumo['A11'].font = Font(bold=True)
            row = 12
            for key, value in filtros_aplicados.items():
                ws_resumo[f'A{row}'] = f"• {key.replace('_', ' ').title()}: {value}"
                row += 1
        
        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 20
        
        # ========== ABA: DADOS DETALHADOS ==========
        ws_dados = wb.create_sheet(title="Dados Detalhados")
        
        # Cabeçalho
        headers = ['ID', 'Nome', 'Descrição', 'Tipo', 'Prioridade', 'Status', 'Prazo Final', 'Cliente', 'Secretaria', 'Solicitante', 'Criado em']
        ws_dados.append(headers)
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws_dados[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados
        for demanda in demandas:
            # Obter relacionamentos
            tipo_nome = demanda.tipo_demanda.nome if demanda.tipo_demanda else 'N/A'
            prioridade_nome = demanda.prioridade.nome if demanda.prioridade else 'N/A'
            status_nome = demanda.status.value.replace('_', ' ').title()
            prazo = demanda.prazo_final.strftime('%d/%m/%Y') if demanda.prazo_final else 'N/A'
            criado_em = demanda.created_at.strftime('%d/%m/%Y %H:%M') if demanda.created_at else 'N/A'
            
            # Obter cliente e secretaria
            cliente_nome = 'N/A'
            secretaria_nome = 'N/A'
            if demanda.secretaria:
                secretaria_nome = demanda.secretaria.nome
                if demanda.secretaria.cliente:
                    cliente_nome = demanda.secretaria.cliente.nome
            
            # Obter solicitante
            solicitante = 'N/A'
            if demanda.usuario:
                solicitante = demanda.usuario.nome_completo or demanda.usuario.username
            
            # Descrição (truncar se muito longa)
            descricao = demanda.descricao[:100] + '...' if demanda.descricao and len(demanda.descricao) > 100 else (demanda.descricao or 'N/A')
            
            ws_dados.append([
                demanda.id,
                demanda.nome,
                descricao,
                tipo_nome,
                prioridade_nome,
                status_nome,
                prazo,
                cliente_nome,
                secretaria_nome,
                solicitante,
                criado_em
            ])
        
        # Formatar células de dados
        data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_fill_grey = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for row_idx, row in enumerate(ws_dados.iter_rows(min_row=2, max_row=ws_dados.max_row), start=2):
            for cell in row:
                cell.border = border
                cell.alignment = data_alignment
                # Alternar cores de fundo
                if row_idx % 2 == 0:
                    cell.fill = data_fill_white
                else:
                    cell.fill = data_fill_grey
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 15,  # ID
            'B': 30,  # Nome
            'C': 40,  # Descrição
            'D': 15,  # Tipo
            'E': 15,  # Prioridade
            'F': 15,  # Status
            'G': 12,  # Prazo Final
            'H': 20,  # Cliente
            'I': 20,  # Secretaria
            'J': 20,  # Solicitante
            'K': 18   # Criado em
        }
        for col, width in column_widths.items():
            ws_dados.column_dimensions[col].width = width
        
        # Congelar primeira linha
        ws_dados.freeze_panes = 'A2'
        
        # Salvar no buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer

